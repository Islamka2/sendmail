# -*- coding: utf-8 -*-
# !/usr/bin/env python
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import qrcode
from PIL import Image
from email import encoders
import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
import imghdr
w = 150
h = 150
size = (w,h)
a = 1
b = ''
wb = xl.load_workbook(filename="123.xlsx")
ws = wb.active
sheet = wb['test']
send_len = int(input("Сколько людей?"))
sender_email = "islamkaawad@gmail.com"

password = ""
mails = ['islamkaawad2@gmail.com', 'ermek_2000@mail.ru', 'kesha000099@gmail.com','medvedpwnz@gmail.com']
message = MIMEMultipart("alternative")
message["Subject"] = "multipart test"
message["From"] = sender_email



html = """\
<html>
  <body>
    <p>Hi,<br>
    <h1> Салам </h1>
       Салам ворам<br>
       has many great tutorials.
    </p>
  </body>
</html>
"""
part = MIMEBase('application', "octet-stream")

part2 = MIMEText(html, "html")

message.attach(part2)

context = ssl.create_default_context()
with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
    for i in range(send_len):
      try:
        message = MIMEMultipart("alternative")
        message["Subject"] = "multipart test"
        message["From"] = sender_email

        # Create the plain-text and HTML version of your message
        text = ""
        html = """\
        <html>
          <body>
            <b>Здравствуйте!</b>
            <p>Благодарим, что Вы с нами.<br>

  Ваше приглашение и программа мероприятия во вложении к данному письму<br>
  <br>
  <br>
  <br>
  Young Entrepreneurs Business Forum/Форум Молодых Предпринимателей – это новая площадка, которая призвана объединить молодых бизнесменов Казахстана, верящих в силу новых идей и возможностей.
  <br><br><br>
  Адрес: ул. Досмухамедова, 115. Отель DoubleTree by Hilton Almaty.
  <br>
  ВХОД ТОЛЬКО ПО ПЕРСОНАЛЬНОМУ QR-КОДУ.
  <br><br><br>
  ВАЖНО!
  УЖИН ПЛАТНЫЙ.
  <br><br>
  Стоимость — 10 000 тенге. Оплата по желанию.
  <br><br>
  Посещение форума не обязывает покупать ваучер на ужин. Вход в ресторан только по ваучерам. МЕСТА ОГРАНИЧЕНЫ.
  <br><br>
  Заявка на получение ваучера будет принята при поступлении платежа на счет. Номер счета выдается после отправления заявки на Whatsapp номер: +7 707 573 16 93
  <br><br>
  Заявка на ужин принимается до 12.00/10.12.19
  <br><br>
  Заявка, отправленная после указанного срока не принимается, место не бронируется.
            </p>
          </body>
        </html>
        """


        part = MIMEBase('application', "octet-stream")

        part2 = MIMEText(html, "html")

        message.attach(part2)
        num = int(sheet.cell(row=i+1,column=1).value)
        name = str(sheet.cell(row=i+1, column=2).value)
        surname = sheet.cell(row=i+1, column=3).value
        email = sheet.cell(row=i+1, column=4).value
        b = str(a)
        a += 1
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=8,
            border=1,
        )
        test_one = "Islam"
        test_two = "Ислам"
        print(name, surname, num)
        text = u'Welcome {0} {1}. Your individual number - {2}.'.format(name, surname, num)

        qr.add_data(text)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")
        img.save("qrcodes/%s.jpg"%b)
        qrcod = Image.open('qrcodes/%s.jpg'%b)
        main_image = Image.open('main.jpg')
        main_image.paste(qrcod,(590,1035))

        main_image.save("mail/mail%s.jpg"% b)

                



        filename = "mail/mail%s.jpg"% b
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(filename, "rb").read())
        part.add_header('Content-Disposition', 'attachment; filename="invite.jpg"')
        encoders.encode_base64(part)
        message.attach(part)
        message["To"] = email
        server.login(sender_email, password)
        server.sendmail(
            sender_email, email, message.as_string()
        )
        print("Сообщение номер {0} по адресу {1} , было доставлено. Имя получателя: {2} , фамилия: {3}".format(num,email,name,surname))
      except:
        print("Некорректный адрес")
        continue
print("Доставка была успешно завершена.")